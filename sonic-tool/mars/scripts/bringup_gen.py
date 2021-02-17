#!/usr/bin/env python
import argparse
import xml.etree.ElementTree as ET
import json
import os
import re
import shutil
import sys
import traceback
from openpyxl import load_workbook
from os import path
from tabulate import tabulate
from xml.etree.ElementTree import Element, SubElement, Comment, ElementTree

class BringupToolException(Exception):
    def __init__(self, *args):
        if args:
            self.message = args[0]
        else:
            self.message = None

    def __str__(self):
        if self.message:
            return 'BringupToolException, {0}'.format(self.message)
        else:
            return 'BringupToolException has been raised'

"""
usage: bringup_gen.py [-h][-v][-i [input_folder]]

Bring up a new platform

Optional arguments:
  -h, --help            Show this help message and exit
  -v, --version         Show program's version number and exit
  -i, --input           input folder that contains all the input files
"""

class BringupGenerator(object):
    """
    Tool for Bringup a platform
    """

    def __init__(self, input_folder, sonic_path):
        self.sonic_path = ""
        self.default_sku = ""
        self.platform = ""
        self.platform_json_file = ""
        self.platform_json_file_exists = False
        self.buffer_spectrum_t0_xlsx_file = ""
        self.buffer_spectrum_t0_xlsx_file_exists = False
        self.buffer_spectrum_t1_xlsx_file = ""
        self.buffer_spectrum_t2_xlsx_file_exists = False
        self.sai_profile_file = ""
        self.sai_profile_file_exists = False
        self.sai_xml_file = ""
        self.sai_xml_file_exists = False
        self.hwsku_json_file = ""
        self.hwsku_json_file_exists = False
        self.default_sku_folder = ""
        self.bringup_summary = []
        self.portconfig_dict = {}
        self.set_input_folder(input_folder)
        self.set_sonic_path_folder(sonic_path)
        self.set_output_folder()

    def set_input_folder(self, input_folder):
        if (not path.exists(input_folder) or path.isfile(input_folder)):
            raise BringupToolException("Input folder ", input_folder, " for Bringup Tool does not exist. Exiting...")
        self.input_folder = input_folder

    def set_sonic_path_folder(self, sonic_path):
        if (not path.exists(sonic_path) or path.isfile(sonic_path)):
            raise BringupToolException("Sonic-buildimage path ", sonic_path, " required for Bringup Tool does not exist. Exiting...")

        pattern = '.*sonic-buildimage/?'
        m = re.match(pattern, sonic_path)
        if m is None:
            raise BringupToolException("Error: sonic-path (" + sonic_path + ") does not contain sonic-buildimage workspace. Exiting...")

        device_mellanox_path = sonic_path + "/device/mellanox/"
        if (not path.exists(device_mellanox_path) or path.isfile(device_mellanox_path)):
            raise BringupToolException("Error: sonic-path (" + sonic_path + ") does not contain /device/mellanox path. Exiting...")

        if not sonic_path.endswith("/"):
            sonic_path = sonic_path + "/"

        self.sonic_path = sonic_path

    def set_output_folder(self):
        platform_json_file = self.input_folder + "/platform.json"
        if path.exists(platform_json_file) and path.isfile(platform_json_file):
            self.platform_json_file = platform_json_file
            self.platform_json_file_exists = True
        else:
            raise BringupToolException("Error: Important Input file platform.json does not exist. Exiting...")

        try:
            with open(self.platform_json_file, 'r') as f:
                json_data = json.load(f)
        except IOError:
            raise BringupToolException("Couldn't open file: " + self.platform_json_file)
        chassis_name = json_data['chassis']['name']
        if not chassis_name.startswith("MSN"):
            raise BringupToolException("Chassis name does not follow format. Exiting...")

        self.platform = chassis_name[3:].lower()
        self.default_sku = "ACS-MSN" + self.platform

        device_mellanox_path = self.sonic_path + "device/mellanox/"
        self.output_folder = device_mellanox_path + "x86_64-mlnx_msn" + self.platform + "-r0/"

    def generate_symbolic_link_file(self, target_file, src_file):
        filename = os.path.basename(target_file)
        output_folder = os.path.dirname(target_file)
        cwd = os.getcwd()
        os.chdir(output_folder)
        if not (path.exists(src_file) and path.isfile(src_file)):
            self.bringup_summary.append(filename + ": File not created - source file missing for symlink")
            os.chdir(cwd)
            return
        else:
             if (path.exists(target_file) and path.isfile(target_file)):
                 os.remove(target_file)
             os.symlink(src_file, target_file)
             os.chdir(cwd)
             self.bringup_summary.append(filename + ": File created")

    def generate_default_sku_file(self):
        default_sku_file = self.output_folder + "default_sku"
        try:
            with open(default_sku_file, 'w') as f:
                f.write(self.default_sku)
        except IOError:
            raise BrintupToolException("Couldn't open file: " + self.default_sku_file)
        self.bringup_summary.append("default_sku: File created")

    def generate_pcie_yaml_file(self):
        pcie_yaml_file = self.output_folder + "pcie.yaml"
        try:
            with open(pcie_yaml_file, 'w') as f:
                f.write("Place-Holder file created. To actually create the file, first remove this file and run \"sudo pcieutil generate\" on the switch")
        except IOError:
            raise BringupToolException("Couldn't open file: " + pcie_yaml_file)
        self.bringup_summary.append("pcie.yaml: Place-Holder file created")

    def generate_platform_json_file(self):
        if not self.platform_json_file_exists:
            self.bringup_summary.append("platform.json: Input file missing")
        else:
            dest_platform_json_file = self.output_folder + "platform.json"
            shutil.copyfile(self.platform_json_file, dest_platform_json_file)
            self.bringup_summary.append("platform.json: File created")

    def generate_platform_components_json_file(self):
        if not self.platform_json_file_exists:
            self.bringup_summary.append("port_config.ini: File not created - Input file missing")
            return

        try:
            with open(self.platform_json_file, 'r') as f:
                json_data = json.load(f)
        except IOError:
            raise BringupToolException("Couldn't open file: " + self.platform_json_file)
        comp_list = json_data['chassis']['components']
        platform_str = "MSN" + self.platform.lower()
        chassis_dict = {
            "chassis" : {
                platform_str : {
                    "component" : {
                    }
                }
            }
        }

        for i in comp_list:
            comp = i['name']
            chassis_dict["chassis"][platform_str]["component"][comp] = {}

        json_object = json.dumps(chassis_dict, sort_keys=True, indent=4)
        platform_components_json_file = self.output_folder + "platform_components.json"
        with open(platform_components_json_file, 'w') as outfile:
            outfile.write(json_object)
        self.bringup_summary.append("platform_components.json: File created")

    def generate_platform_reboot_file(self):
        self.generate_symbolic_link_file(self.output_folder + "platform_reboot", "../x86_64-mlnx_msn2700-r0/" + "platform_reboot")

    def generate_platform_wait_file(self):
        self.generate_symbolic_link_file(self.output_folder + "platform_wait", "../x86_64-mlnx_msn2700-r0/" + "platform_wait")

    def generate_pmon_daemon_control_json_file(self):
        self.generate_symbolic_link_file(self.output_folder + "pmon_daemon_control.json", "../x86_64-mlnx_msn2700-r0/" + "pmon_daemon_control.json")

    def generate_port_peripher_config_j2_file(self):
        wb = load_workbook(self.buffer_spectrum_t0_xlsx_file, data_only=True)
        headroom_ws = wb["headroom"]
        gearbox_type = headroom_ws['U2'].value

        if gearbox_type == "Non":
            self.bringup_summary.append("port_peripher_config.j2: File not required for this platform")
            return

        port_peripher_config_j2_file = self.output_folder + "port_peripher_config.j2"
        file_contents_str = """[
    {{%- include peripheral_table.j2 %}},
    {{
        "PORT_PERIPHERAL_TABLE:global": {{
            "gearbox_model": "{gbtype}"
        }},
        "OP": "SET"
    }}
]""".format(gbtype=gearbox_type)
        try:
            with open(port_peripher_config_j2_file, 'w') as f:
                f.write(file_contents_str)
        except IOError:
            raise BringupToolException("Couldn't open file: " + port_peripher_config_j2_file)
        self.bringup_summary.append("port_peripher_config.j2.json: File created")

    def generate_sensors_conf_file(self):
        self.bringup_summary.append("sensors.conf: File not created - File is not automatable")

    def generate_system_health_monitoring_config_json_file(self):
        self.generate_symbolic_link_file(self.output_folder + "system_health_monitoring_config.json", "../x86_64-mlnx_msn2700-r0/" + "system_health_monitoring_config.json")

    def generate_thermal_policy_json_file(self):
        self.generate_symbolic_link_file(self.output_folder + "thermal_policy.json", "../x86_64-mlnx_msn2700-r0/" + "thermal_policy.json")

    def generate_eeprom_py_file(self):
        self.generate_symbolic_link_file(self.output_folder + "plugins/eeprom.py", "../../x86_64-mlnx_msn2700-r0/plugins/" + "eeprom.py")

    def generate_fanutil_py_file(self):
        self.generate_symbolic_link_file(self.output_folder + "plugins/fanutil.py", "../../x86_64-mlnx_msn2700-r0/plugins/" + "fanutil.py")

    def generate_psuutil_py_file(self):
        self.generate_symbolic_link_file(self.output_folder + "plugins/psuutil.py", "../../x86_64-mlnx_msn2700-r0/plugins/" + "psuutil.py")

    def generate_sfplpmget_py_file(self):
        self.generate_symbolic_link_file(self.output_folder + "plugins/sfplpmget.py", "../../x86_64-mlnx_msn2700-r0/plugins/" + "sfplpmget.py")

    def generate_sfplpmset_py_file(self):
        self.generate_symbolic_link_file(self.output_folder + "plugins/sfplpmset.py", "../../x86_64-mlnx_msn2700-r0/plugins/" + "sfplpmset.py")

    def generate_sfpreset_py_file(self):
        self.generate_symbolic_link_file(self.output_folder + "plugins/sfpreset.py", "../../x86_64-mlnx_msn2700-r0/plugins/" + "sfpreset.py")

    def generate_sfputil_py_file(self):
        self.generate_symbolic_link_file(self.output_folder + "plugins/sfputil.py", "../../x86_64-mlnx_msn2700-r0/plugins/" + "sfputil.py")

    def generate_buffers_json_j2_file(self):
        self.generate_symbolic_link_file(self.default_sku_folder + "buffers.json.j2", "../../x86_64-mlnx_msn2700-r0/ACS-MSN2700/" + "buffers.json.j2")

    def generate_buffers_defaults_j2_file(self, topo):
        if topo == 't0':
            if not self.buffer_spectrum_t0_xlsx_file_exists:
                self.bringup_summary.append("buffers_defaults_t0.j2: File not created - Input file missing")
                return
        else:
            if not self.buffer_spectrum_t1_xlsx_file_exists:
                self.bringup_summary.append("buffers_defaults_t1.j2: File not created - Input file missing")
                return

        if topo == 't0':
            wb = load_workbook(self.buffer_spectrum_t0_xlsx_file, data_only=True)
        else:
            wb = load_workbook(self.buffer_spectrum_t1_xlsx_file, data_only=True)

        sonic_ms_ws = wb["SoNIC MS "]
        ing_lossless_value = sonic_ms_ws['M3'].value
        ing_lossy_value = sonic_ms_ws['M3'].value
        egr_lossless_value =  sonic_ms_ws['M4'].value
        egr_lossy_value =  sonic_ms_ws['M5'].value

        if egr_lossless_value.lower().find("infinite") != -1:
            egr_lossless_value = sonic_ms_ws['P16'].value
        else:
            if egr_lossless_value.isdigit():
                egr_lossless_value = int(egr_lossless_value)
            else:
                egr_lossless_value = 0

        if not self.default_sku:
            if topo == 't0':
                self.bringup_summary.append("buffers_defaults_t0.j2: File not created - Input info missing")
                return
            else:
                self.bringup_summary.append("buffers_defaults_t1.j2: File not created - Input info missing")
                return
        else:
            if not self.default_sku.startswith("Mellanox"):
                #SKU is non MSFT SKU
                ing_lossless_pool_size = ing_lossless_value * 512
                ing_lossy_pool_size = ing_lossy_value * 512
                egr_lossless_pool_size = egr_lossless_value * 512
                egr_lossy_pool_size = egr_lossy_value * 512
            else:
                #SKU is MSFT SKU
                ing_lossless_pool_size = ing_lossless_value * 1024
                ing_lossy_pool_size = ing_lossy_value * 1024
                egr_lossless_pool_size = egr_lossless_value * 1024
                egr_lossy_pool_size = egr_lossy_value * 1024

        if topo == 't0':
            buffers_defaults_j2_file = self.default_sku_folder + "buffers_defaults_t0.j2"
            buffers_defaults_j2_filename = "buffers_defaults_t0.j2"
        else:
            buffers_defaults_j2_file = self.default_sku_folder + "buffers_defaults_t1.j2"
            buffers_defaults_j2_filename = "buffers_defaults_t1.j2"
        file_contents_str = """{{% set default_cable = '5m' %}}
{{% set ingress_lossless_pool_size =  '{illps}' %}}
{{% set ingress_lossy_pool_size =  '{ilyps}' %}}
{{% set egress_lossless_pool_size =  '{ellps}' %}}
{{% set egress_lossy_pool_size =  '{elyps}' %}}

{{%- macro generate_port_lists(PORT_ALL) %}}
    {{# Generate list of ports #}}
    {{%- for port_idx in range(0, 32) %}}
        {{%- if PORT_ALL.append("Ethernet%d" % (port_idx)) %}}{{%- endif %}}
    {{%- endfor %}}
{{%- endmacro %}}

{{%- macro generate_buffer_pool_and_profiles() %}}
    "BUFFER_POOL": {{
        "ingress_lossless_pool": {{
            {{%- if dynamic_mode is not defined %}}
            "size": "{{{{ ingress_lossless_pool_size }}}}",
            {{%- endif %}}
            "type": "ingress",
            "mode": "dynamic"
        }},
        "ingress_lossy_pool": {{
            {{%- if dynamic_mode is not defined %}}
            "size": "{{{{ ingress_lossy_pool_size }}}}",
            {{%- endif %}}
            "type": "ingress",
            "mode": "dynamic"
        }},
        "egress_lossless_pool": {{
            "size": "{{{{ egress_lossless_pool_size }}}}",
            "type": "egress",
            "mode": "dynamic"
        }},
        "egress_lossy_pool": {{
            {{%- if dynamic_mode is not defined %}}
            "size": "{{{{ egress_lossy_pool_size }}}}",
            {{%- endif %}}
            "type": "egress",
            "mode": "dynamic"
        }}
    }},
    "BUFFER_PROFILE": {{
        "ingress_lossless_profile": {{
            "pool":"[BUFFER_POOL|ingress_lossless_pool]",
            "size":"0",
            "dynamic_th":"7"
        }},
        "ingress_lossy_profile": {{
            "pool":"[BUFFER_POOL|ingress_lossy_pool]",
            "size":"0",
            "dynamic_th":"3"
        }},
        "egress_lossless_profile": {{
            "pool":"[BUFFER_POOL|egress_lossless_pool]",
            "size":"0",
            "dynamic_th":"7"
        }},
        "egress_lossy_profile": {{
            "pool":"[BUFFER_POOL|egress_lossy_pool]",
            "size":"9216",
            "dynamic_th":"7"
        }},
        "q_lossy_profile": {{
            "pool":"[BUFFER_POOL|egress_lossy_pool]",
            "size":"0",
            "dynamic_th":"3"
        }}
    }},
{{%- endmacro %}}

{{%- macro generate_profile_lists(port_names) %}}
    "BUFFER_PORT_INGRESS_PROFILE_LIST": {{
{{% for port in port_names.split(',') %}}
        "{{{{ port }}}}": {{
            "profile_list" : "[BUFFER_PROFILE|ingress_lossless_profile],[BUFFER_PROFILE|ingress_lossy_profile]"
        }}{{% if not loop.last %}},{{% endif %}}

{{% endfor %}}
    }},
    "BUFFER_PORT_EGRESS_PROFILE_LIST": {{
{{% for port in port_names.split(',') %}}
        "{{{{ port }}}}": {{
            "profile_list" : "[BUFFER_PROFILE|egress_lossless_profile],[BUFFER_PROFILE|egress_lossy_profile]"
        }}{{% if not loop.last %}},{{% endif %}}

{{% endfor %}}
    }}
{{%- endmacro %}}

{{%- macro generate_queue_buffers(port_names) %}}
    "BUFFER_QUEUE": {{
{{% for port in port_names.split(',') %}}
        "{{{{ port }}}}|3-4": {{
            "profile" : "[BUFFER_PROFILE|egress_lossless_profile]"
        }},
{{% endfor %}}
{{% for port in port_names.split(',') %}}
        "{{{{ port }}}}|0-2": {{
            "profile" : "[BUFFER_PROFILE|q_lossy_profile]"
        }},
{{% endfor %}}
{{% for port in port_names.split(',') %}}
        "{{{{ port }}}}|5-6": {{
            "profile" : "[BUFFER_PROFILE|q_lossy_profile]"
        }}{{% if not loop.last %}},{{% endif %}}

{{% endfor %}}
    }}
{{%- endmacro %}}""".format(illps=ing_lossless_pool_size, ilyps=ing_lossy_pool_size, ellps=egr_lossless_pool_size, elyps=egr_lossy_pool_size)
        try:
            with open(buffers_defaults_j2_file, 'w') as f:
                f.write(file_contents_str)
        except IOError:
            raise BringupToolException("Couldn't open file: " + buffers_defaults_j2_file)
        self.bringup_summary.append(buffers_defaults_j2_filename + ": File created")

    def generate_buffers_dynamic_json_j2_file(self):
        self.generate_symbolic_link_file(self.default_sku_folder + "buffers_dynamic.json.j2", "../../x86_64-mlnx_msn2700-r0/ACS-MSN2700/" + "buffers_dynamic.json.j2")

    def generate_hwsku_json_file(self):
        if not self.hwsku_json_file_exists:
            self.bringup_summary.append("hwsku.json: Input file missing")
        else:
            dest_hwsku_json_file = self.default_sku_folder + "hwsku.json"
            shutil.copyfile(self.hwsku_json_file, dest_hwsku_json_file)
            self.bringup_summary.append("hwsku.json: File created")

    def generate_pg_profile_lookup_ini_file(self):
        if not self.buffer_spectrum_t0_xlsx_file_exists:
            self.bringup_summary.append("pg_profile_lookup.ini: File not created - Input file missing")
            return

        wb = load_workbook(self.buffer_spectrum_t0_xlsx_file, data_only=True)
        headroom_ws = wb["headroom"]

        pg_profile_header = ["# PG lossless profiles"]
        pg_profile = []

        index = 17
        while True:
            ae_index = 'AE' + str(index)
            af_index = 'AF' + str(index)
            ag_index = 'AG' + str(index)
            ah_index = 'AH' + str(index)
            ai_index = 'AI' + str(index)
            af_str = str(headroom_ws[af_index].value) + 'm'
            pg_list = [headroom_ws[ae_index].value, af_str, headroom_ws[ag_index].value, headroom_ws[ah_index].value, headroom_ws[ai_index].value]
            if not headroom_ws[ae_index].value:
                break
            pg_profile.append(pg_list)
            index = index + 1

        pg_profile_lookup_ini_file = self.default_sku_folder + "pg_profile_lookup.ini"
        try:
            with open(pg_profile_lookup_ini_file, 'w') as pg_file:
                pg_file.write("# PG lossless profiles\n")
                pg_profile.sort(key=lambda x: (int(x[1][:-1]), x[0])) # sort the list with cable length and speed
                pg_file.write(tabulate(pg_profile, tablefmt="plain"))
        except IOError:
            raise BringupToolException("Couldn't open file: " + pg_profile_lookup_ini_file)
        self.bringup_summary.append("pg_profile_lookup.ini: File created")

    def generate_port_config_ini_file(self):

        if not self.platform_json_file_exists:
            self.bringup_summary.append("port_config.ini: File not created - Input file missing")
            return

        try:
            with open(self.platform_json_file, 'r') as f:
                json_data = json.load(f)
        except IOError:
            raise BringupToolException("Couldn't open file: " + self.platform_json_file)
        for ifname in json_data['interfaces']:
            if_dict = json_data['interfaces'][ifname]
            idx_str = if_dict['index']
            idx_list = idx_str.split(',')
            orig_index = idx_list[0]
            lanes_str = if_dict['lanes']
            alias_str = if_dict['alias_at_lanes']
            alias_list = alias_str.split(',')
            orig_alias_str = alias_list[0][:-1]

            self.portconfig_dict[orig_index] = [ifname, lanes_str, orig_alias_str, str(orig_index)]

        portconfig_header = ["# name", "lanes", "alias", "index"]
        port_config = []
        for line in self.portconfig_dict.values():
            port_config.append(line)
        port_config.sort(key=lambda x: int(x[3])) # sort the list with interface name
        port_config_ini_file = self.default_sku_folder + "port_config.ini"
        try:
            with open(port_config_ini_file,"w") as pcini_file:
                pcini_file.write(tabulate(port_config, portconfig_header, tablefmt="plain"))
        except IOError:
            raise BringupToolException("Couldn't open file: " + port_config_ini_file)

    def generate_qos_json_j2_file(self):
        self.generate_symbolic_link_file(self.default_sku_folder + "qos.json.j2", "../../x86_64-mlnx_msn2700-r0/ACS-MSN2700/" + "qos.json.j2")

    def generate_sai_profile_file(self):
        if not self.sai_profile_file_exists:
            self.bringup_summary.append("sai.profile: Input file missing")
        else:
            dest_sai_profile_file = self.default_sku_folder + "sai.profile"
            shutil.copyfile(self.sai_profile_file, dest_sai_profile_file)
            self.bringup_summary.append("sai.profile: File created")

    def generate_sai_xml_file(self):
        sai_xml_str = "sai_" + self.platform.lower() + ".xml"
        if not self.sai_xml_file_exists:
            self.bringup_summary.append(sai_xml_str + ": Input file missing")
        else:
            dest_sai_xml_file = self.default_sku_folder + sai_xml_str
            shutil.copyfile(self.sai_xml_file, dest_sai_xml_file)
            self.bringup_summary.append(sai_xml_str + ": File created")

    def generate_platform_files(self):
        self.generate_default_sku_file()
        self.generate_pcie_yaml_file()
        self.generate_platform_json_file()
        self.generate_platform_components_json_file()
        self.generate_platform_reboot_file()
        self.generate_platform_wait_file()
        self.generate_pmon_daemon_control_json_file()
        self.generate_port_peripher_config_j2_file()
        self.generate_sensors_conf_file()
        self.generate_system_health_monitoring_config_json_file()
        self.generate_thermal_policy_json_file()

    def generate_plugin_files(self):
        self.generate_eeprom_py_file()
        self.generate_fanutil_py_file()
        self.generate_psuutil_py_file()
        self.generate_sfplpmget_py_file()
        self.generate_sfplpmset_py_file()
        self.generate_sfpreset_py_file()
        self.generate_sfputil_py_file()

    def generate_sku_files(self):
        self.generate_buffers_json_j2_file()
        self.generate_buffers_defaults_j2_file('t0')
        self.generate_buffers_defaults_j2_file('t1')
        self.generate_buffers_dynamic_json_j2_file()
        self.generate_hwsku_json_file()
        self.generate_pg_profile_lookup_ini_file()
        self.generate_port_config_ini_file()
        self.generate_qos_json_j2_file()
        self.generate_sai_profile_file()
        self.generate_sai_xml_file()

    def display_bringup_summary(self):
        print("\n\nBringup Generation Tool Summary: ")
        print("----------------------------------")
        for bu_str in self.bringup_summary:
            print(bu_str)
        print("----------------------------------")

    def create_output_folders(self):
        #Create Output directory
        print("Bringup Tool Output folder is: " + self.output_folder)
        if not os.path.exists(self.output_folder):
            os.mkdir(self.output_folder)

        plugins_folder = self.output_folder + "plugins"
        if not os.path.exists(plugins_folder):
            os.mkdir(plugins_folder)

        self.default_sku_folder = self.output_folder + self.default_sku + "/"
        if not os.path.exists(self.default_sku_folder):
            os.mkdir(self.default_sku_folder)

    def check_input_files(self):
        buffer_spectrum_t0_xlsx_file = self.input_folder + "/buffer_spectrum_t0.xlsx"
        if (path.exists(buffer_spectrum_t0_xlsx_file) and path.isfile(buffer_spectrum_t0_xlsx_file)):
            self.buffer_spectrum_t0_xlsx_file = buffer_spectrum_t0_xlsx_file
            self.buffer_spectrum_t0_xlsx_file_exists = True
        else:
            print("WARN: Input file spectrum_buffer_t0.xlsx does not exist.")

        buffer_spectrum_t1_xlsx_file = self.input_folder + "/buffer_spectrum_t1.xlsx"
        if (path.exists(buffer_spectrum_t1_xlsx_file) and path.isfile(buffer_spectrum_t1_xlsx_file)):
            self.buffer_spectrum_t1_xlsx_file = buffer_spectrum_t1_xlsx_file
            self.buffer_spectrum_t1_xlsx_file_exists = True
        else:
            print("WARN: Input file spectrum_buffer_t1.xlsx does not exist.")

        sai_profile_file = self.input_folder + "/sai.profile"
        if (path.exists(sai_profile_file) and path.isfile(sai_profile_file)):
            self.sai_profile_file = sai_profile_file
            self.sai_profile_file_exists = True
        else:
            print("WARN: Input file sai.profile does not exist.")

        sai_xml_str = "sai_" + self.platform.lower() + ".xml"
        sai_xml_file = self.input_folder + "/" + sai_xml_str
        if (path.exists(sai_xml_file) and path.isfile(sai_xml_file)):
            self.sai_xml_file = sai_xml_file
            self.sai_xml_file_exists = True
        else:
            print("WARN: Input file " + sai_xml_str + " does not exist.")

        hwsku_json_file = self.input_folder + "/hwsku.json"
        if (path.exists(hwsku_json_file) and path.isfile(hwsku_json_file)):
            self.hwsku_json_file = hwsku_json_file
            self.hwsku_json_file_exists = True
        else:
            print("WARN: Input file hwsku.json does not exist.")

def main(argv):
    parser = argparse.ArgumentParser(description='Generate files for Bringing up a platform. The 2 inputs for the tool are input folder - where all the input files exist and sonic-path where the sonic-buildimage workspace exists and the files will be generated under /device/mellanox/ folder of sonic-buildimage workspace. \n\nThe important input files are platform.json, buffers_spectrum_t0.xlsx and buffers_spectrum_t1.xlsx. These files have all the required information required to generate the platform files. Other input files are hwsku.json, sai.profile and sai xml file. In the buffer_spectrum xlsx files, please set the correct ASIC type, gearbox models and the number of ports per headroom size.',
                                      formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument('-v', '--version', action='version', version='%(prog)s 1.0')
    parser.add_argument('-i', '--input', action='store', nargs=1, required=True, help='Input folder containing files that are input for Bringup tool', default=None)
    parser.add_argument('-p', '--sonicpath', action='store', nargs=1, required=True, help='path of sonic-buildimage workspace', default=None)

    args=parser.parse_args()

    try:
        bringup = BringupGenerator(args.input[0], args.sonicpath[0])
        bringup.create_output_folders()
        bringup.check_input_files()
        bringup.generate_platform_files()
        bringup.generate_plugin_files()
        bringup.generate_sku_files()
        bringup.display_bringup_summary()

    except Exception:
        traceback.print_exc(file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main(sys.argv)
